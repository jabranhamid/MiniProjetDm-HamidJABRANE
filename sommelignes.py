import pandas as pd
from openpyxl import load_workbook

def calculer_somme(df, plages_colonnes):
    colonnes_calculees = pd.DataFrame()  # DataFrame pour stocker les colonnes calculées

    for plage_colonnes in plages_colonnes:
        somme_colonne = df.iloc[:, plage_colonnes].sum(axis=1)
        nom_colonne = f'Somme{plage_colonnes[0]}_{plage_colonnes[-1]}'  # Nom de la colonne basé sur les indices de la plage
        colonnes_calculees[nom_colonne] = somme_colonne

    return colonnes_calculees

# Charger les données depuis un fichier Excel
file_path = 'EstimeSoi_IA2.xlsx'
df = pd.read_excel(file_path)

# Spécifier les plages de colonnes pour les différents calculs
plages_colonnes_test1 = list(range(13, 23))
plages_colonnes_test2 = list(range(23, 44))

# Appeler la fonction pour calculer les sommes pour différentes plages de colonnes
df_sommes = calculer_somme(df, [plages_colonnes_test1, plages_colonnes_test2])

# Charger le fichier Excel destination
nouveau_file_path = 'excel2.xlsx'
book = load_workbook(nouveau_file_path)

# Sélectionner la première feuille du fichier Excel destination
feuille_destination = book.active

feuille_destination.cell(row=1, column=3, value='TestEstimeDeSoi')
feuille_destination.cell(row=1, column=4, value='TestDepression')

# Ajouter les valeurs calculées à différentes colonnes
for index, row in df_sommes.iterrows():
    for col_index, value in enumerate(row, start=3):
        feuille_destination.cell(row=index + 2, column=col_index, value=value)

# Enregistrer les modifications dans le fichier Excel destination
book.save(nouveau_file_path)

print("Les valeurs des sommes ont été ajoutées avec succès dans 'excel2.xlsx'.")
####--------------------
# Charger le fichier Excel source
df_source = pd.read_excel(file_path)

# Extraire la colonne 'Age'
colonne_age = df_source['Age']

# Charger le fichier Excel destination
book = load_workbook(nouveau_file_path)

# Sélectionner la première feuille du fichier Excel destination
feuille_destination = book.active
# Ajouter l'en-tête 'Age' à la première ligne de la première colonne
feuille_destination.cell(row=1, column=1, value='Age')

# Ajouter les valeurs de la colonne 'Age' à la première colonne
for index, value in enumerate(colonne_age, start=2):
    feuille_destination.cell(row=index, column=1, value=value)

# Enregistrer les modifications dans le fichier Excel destination
book.save(nouveau_file_path)
print("Les valeurs de la colonne 'Age' ont été copiées avec succès dans 'excel2.xlsx'.")
