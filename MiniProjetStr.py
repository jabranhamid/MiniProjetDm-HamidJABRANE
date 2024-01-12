# -*- coding: utf-8 -*-
"""
Created on Mon Nov 20 11:12:29 2023

@author: Hamid Jabrane
"""
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import cv2
import matplotlib.pyplot as plt

import numpy as np
from PIL import Image
import os
from scipy.cluster.vq import vq
import joblib
import seaborn as sns
from sklearn.metrics import confusion_matrix, accuracy_score

# Affichage de l'image logo de FSTE et MASTER SIDI 
image_path = "enteteLogoIA.jpg"
st.image(image_path, use_column_width=True)
# le titre et le sujet de l'application Mini Projet
st.sidebar.write("<h2>MINI PROJET <font color='#01F5FA'>IA</font> </h2>", unsafe_allow_html=True)

st.sidebar.markdown(
             
    "<h3 style='color: blue;'>Encadré par: Pr. Imad ZEROUAL</h3>"
    "<h3 style='color: #040290;'>Réalisé par : Hamid JABRANE</h3> "
        "<h3>     </h3> " 
        
        "<h1></h1>"
        "<h4> Entrainer le modèle </h4>"
        "<h4>Tester le modèle</h4>"
        "<h4>Faire la prédiction</h4>"
    , unsafe_allow_html=True
    
    )


def import_excel():
    st.title("Importer un fichier Excel avec Streamlit")

    # Ajouter un bouton pour permettre à l'utilisateur de télécharger le fichier Excel
    uploaded_file = st.file_uploader("Choisissez un fichier Excel", type=["xlsx", "xls"])

    # Vérifier si un fichier a été téléchargé
    if uploaded_file is not None:
        return pd.read_excel(uploaded_file)
    else:
        return None

def afficher_donnees(df):
    st.subheader("Les premières données du fichier Excel:")
    st.dataframe(df.head())

def afficher_taille(df):
    st.subheader("Taille du fichier:")
    st.write(f"Nombre de lignes: {df.shape[0]}")
    st.write(f"Nombre de colonnes: {df.shape[1]}")
def detecter_cellules_vides(df):
    cellules_vides = df.applymap(lambda x: pd.isna(x) or x == "")
    nb_valeurs_manquantes = cellules_vides.sum().sum()
    return cellules_vides, nb_valeurs_manquantes
def calculer_y(df, colonnes):
    # Calculer la somme des colonnes spécifiées
    somme_colonnes = df[colonnes].sum(axis=1)
    
    # Appliquer la formule y=(1/somme)*10
    y = (1 / somme_colonnes) * 100
    
    return y
def calculer_somme(df, plages_colonnes):
    colonnes_calculees = pd.DataFrame()  # DataFrame pour stocker les colonnes calculées

    for plage_colonnes in plages_colonnes:
        somme_colonne = df.iloc[:, plage_colonnes].sum(axis=1)
        nom_colonne = f'Somme{plage_colonnes[0]}_{plage_colonnes[-1]}'  # Nom de la colonne basé sur les indices de la plage
        colonnes_calculees[nom_colonne] = somme_colonne

    return colonnes_calculees
def afficher_info_excel(fichier_excel):
    st.subheader(f"Informations sur le nouveau fichier Excel '{fichier_excel}' :")

    # Charger le fichier Excel
    df_excel = pd.read_excel(fichier_excel)

    # Afficher les premières lignes
    st.write("Aperçu des premières lignes du fichier Excel :")
    st.dataframe(df_excel.head())

    # Afficher la taille du fichier Excel
    st.write(f"Nombre total de lignes : {df_excel.shape[0]}")
    st.write(f"Nombre total de colonnes : {df_excel.shape[1]}")

    # Afficher les noms des colonnes
    st.write("Noms des colonnes :")
    st.write(list(df_excel.columns))

def main():
    # Importer le fichier Excel
    df = import_excel()

    # Vérifier si un fichier a été importé
    if df is not None:
        # Afficher les données et la taille du fichier
        afficher_donnees(df)
        afficher_taille(df)
        # Détecter les cellules vides
        cellules_vides, nb_valeurs_manquantes = detecter_cellules_vides(df)

        # Afficher les cellules vides
        st.subheader("Cellules vides dans la base de données :")
        st.write(cellules_vides)

       # Afficher le nombre de valeurs manquantes
        st.subheader(f"Nombre total de valeurs manquantes : {nb_valeurs_manquantes}")
        # Spécifier les indices de début et de fin de la plage de colonnes
        debut_colonne = 2
        fin_colonne = 12
        # Filtrer uniquement les colonnes numériques
        colonnes_numeriques = df.select_dtypes(include='number').columns

        # Utiliser iloc pour sélectionner la plage de colonnes
        colonnes_pour_y = list(colonnes_numeriques[debut_colonne:fin_colonne+1])

        # Appeler la fonction pour calculer y
        df['Y'] = calculer_y(df, colonnes_pour_y)



        # Charger le fichier Excel destination
        nouveau_file_path = 'DataReduite.xlsx'
        book = load_workbook(nouveau_file_path)

        # Sélectionner la première feuille du fichier Excel destination
        feuille_destination = book.active

        # Ajouter l'en-tête pour la colonne Y
        feuille_destination.cell(row=1, column=2, value='Indicateur_Sociale')

        # Ajouter les valeurs de la colonne Y à la colonne 4 (en commençant à la deuxième ligne)
        for index, value in enumerate(df['Y'], start=1):
            feuille_destination.cell(row=index + 1, column=2, value=value)

        # Enregistrer les modifications dans le fichier Excel destination
        book.save(nouveau_file_path)
        # Spécifier les plages de colonnes pour les différents calculs
        plages_colonnes_test1 = list(range(13, 23))
        plages_colonnes_test2 = list(range(23, 44))

        # Appeler la fonction pour calculer les sommes pour différentes plages de colonnes
        df_sommes = calculer_somme(df, [plages_colonnes_test1, plages_colonnes_test2])

        # Charger le fichier Excel destination
        nouveau_file_path = 'DataReduite.xlsx'
        book = load_workbook(nouveau_file_path)

        # Sélectionner la première feuille du fichier Excel destination
        feuille_destination = book.active

        feuille_destination.cell(row=1, column=3, value='TestEstimeDeSoi')
        feuille_destination.cell(row=1, column=4, value='TestDepression')

        # Ajouter les valeurs calculées à différentes colonnes
        for index, row in df_sommes.iterrows():
            for col_index, value in enumerate(row, start=3):
                feuille_destination.cell(row=index + 2, column=col_index, value=value)

        # Enregistrer les modifications dans le fichier Excel destination
        book.save(nouveau_file_path)

        st.write("Les valeurs des sommes ont été ajoutées avec succès dans 'excel2.xlsx'.")
        ####--------------------
       

        # Extraire la colonne 'Age'
        colonne_age = df['Age']

        # Charger le fichier Excel destination
        book = load_workbook(nouveau_file_path)

        # Sélectionner la première feuille du fichier Excel destination
        feuille_destination = book.active
        # Ajouter l'en-tête 'Age' à la première ligne de la première colonne
        feuille_destination.cell(row=1, column=1, value='Age')

        # Ajouter les valeurs de la colonne 'Age' à la première colonne
        for index, value in enumerate(colonne_age, start=2):
            feuille_destination.cell(row=index, column=1, value=value)

        # Enregistrer les modifications dans le fichier Excel destination
        book.save(nouveau_file_path)
        # Afficher les données et la taille du fichier
        # Utilisez cette fonction dans votre script Streamlit
        st.write("La réduction des dimensions de votre base a été fait avec succès")
        afficher_info_excel('DataReduite.xlsx')
if __name__ == "__main__":
    main()
