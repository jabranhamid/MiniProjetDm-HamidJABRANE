import pandas as pd
from openpyxl import load_workbook

def calculer_y(df, colonnes):
    # Calculer la somme des colonnes spécifiées
    somme_colonnes = df[colonnes].sum(axis=1)
    
    # Appliquer la formule y=(1/somme)*10
    y = (1 / somme_colonnes) * 100
    
    return y

# Charger les données depuis un fichier Excel
file_path = 'EstimeSoi_IA2.xlsx'
df = pd.read_excel(file_path)

# Spécifier les indices de début et de fin de la plage de colonnes
debut_colonne = 2
fin_colonne = 12
# Filtrer uniquement les colonnes numériques
colonnes_numeriques = df.select_dtypes(include='number').columns

# Utiliser iloc pour sélectionner la plage de colonnes
colonnes_pour_y = list(colonnes_numeriques[debut_colonne:fin_colonne+1])

# Appeler la fonction pour calculer y
df['Y'] = calculer_y(df, colonnes_pour_y)



# Charger le fichier Excel destination
nouveau_file_path = 'excel2.xlsx'
book = load_workbook(nouveau_file_path)

# Sélectionner la première feuille du fichier Excel destination
feuille_destination = book.active

# Ajouter l'en-tête pour la colonne Y
feuille_destination.cell(row=1, column=2, value='Indicateur_Sociale')

# Ajouter les valeurs de la colonne Y à la colonne 4 (en commençant à la deuxième ligne)
for index, value in enumerate(df['Y'], start=1):
    feuille_destination.cell(row=index + 1, column=2, value=value)

# Enregistrer les modifications dans le fichier Excel destination
book.save(nouveau_file_path)

print("Les valeurs de la colonne Y ont été ajoutées avec succès dans 'excel2.xlsx'.")
